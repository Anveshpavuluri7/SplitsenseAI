"""
Export service — generates Excel reports using openpyxl.
"""

import io
import uuid
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.receipt import Receipt
from app.models.item import Item
from app.models.settlement import Settlement

logger = logging.getLogger("splitsenseai.services.export")

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class ExportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def generate_excel(self, user_id: uuid.UUID, group_id: uuid.UUID = None) -> bytes:
        wb = Workbook()

        # Sheet 1: Transactions
        ws1 = wb.active
        ws1.title = "Transactions"
        headers = ["Date", "Store", "Item", "Category", "Price", "Qty", "Total"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        query = select(Receipt).where(Receipt.uploaded_by == user_id).order_by(Receipt.receipt_date.desc())
        result = await self.db.execute(query)
        receipts = result.scalars().all()

        row = 2
        for receipt in receipts:
            for item in receipt.items:
                ws1.cell(row=row, column=1, value=str(receipt.receipt_date or "")).border = BORDER
                ws1.cell(row=row, column=2, value=receipt.store_name or "").border = BORDER
                ws1.cell(row=row, column=3, value=item.name).border = BORDER
                ws1.cell(row=row, column=4, value=item.category or "").border = BORDER
                ws1.cell(row=row, column=5, value=float(item.price)).border = BORDER
                ws1.cell(row=row, column=6, value=item.quantity).border = BORDER
                ws1.cell(row=row, column=7, value=float(item.price) * item.quantity).border = BORDER
                row += 1

        for col in range(1, 8):
            ws1.column_dimensions[chr(64 + col)].width = 18

        # Sheet 2: Settlements
        ws2 = wb.create_sheet("Settlements")
        s_headers = ["Date", "From", "To", "Amount", "Status"]
        for col, h in enumerate(s_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER

        if group_id:
            s_result = await self.db.execute(
                select(Settlement).where(Settlement.group_id == group_id).order_by(Settlement.settled_at.desc())
            )
            row = 2
            for s in s_result.scalars().all():
                ws2.cell(row=row, column=1, value=str(s.settled_at or "")).border = BORDER
                ws2.cell(row=row, column=2, value=str(s.from_user)).border = BORDER
                ws2.cell(row=row, column=3, value=str(s.to_user)).border = BORDER
                ws2.cell(row=row, column=4, value=float(s.amount)).border = BORDER
                ws2.cell(row=row, column=5, value=s.status).border = BORDER
                row += 1

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
